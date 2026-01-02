from flask import Flask, render_template, request, send_file, redirect, url_for, flash, session
import os, signal

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime
from io import BytesIO
from openpyxl.styles import Alignment

app = Flask(__name__)
app.secret_key = "gustitos-secret"

# ---------------- CONFIG ----------------
EXCEL_FILE = "plantilla_base.xlsx"
HOJAS_NECESARIAS = [
    "planilla transacciones",
    "planilla repartos",
    "planilla egresos",
    "planilla mermas",
    "planilla desgloses",
    "parametros",
    "planilla cortesias",
]
CIERRES_DIR = "cierres"
os.makedirs(CIERRES_DIR, exist_ok=True)

MEDIOS_VALIDOS = [
    "efectivo", "debito", "credito", "transferencia",
    "pluxee", "edenred", "amipass", "pedidos ya","uber eats"
]
DENOMINACIONES = [10, 50, 100, 500, 1000, 2000, 5000, 10000, 20000]

# --------- Filtro de dinero para Jinja ---------
@app.template_filter("money")
def money(value):
    """Formatea como CLP: $12.345"""
    try:
        if value is None or value == "":
            return "$0"
        value = float(value)
        return f"${value:,.0f}"
    except Exception:
        return "$0"

# -------------- ESTILOS --------------
def _estilos_basicos():
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    return thin_border, header_fill

def _estilizar_encabezado(cells, header_fill, thin_border):
    for c in cells:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.border = thin_border

def _autoajustar_columnas(ws):
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3

def _estilizar_hoja_detalle(ws, header_fill, thin_border):
    if ws.max_row >= 1:
        _estilizar_encabezado(ws[1], header_fill, thin_border)
    if ws.max_row >= 2:
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '"$"#,##0'
    _autoajustar_columnas(ws)

    # ⬇️ Aquí colocas tu nueva función
def resumen_boletas_en_transacciones(wb):
    if "planilla transacciones" not in wb.sheetnames:
        return
    ws = wb["planilla transacciones"]

    thin_border, header_fill = _estilos_basicos()

    # Agrupar por Nº Interno
    boletas = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[2]:
            continue
        nro = str(row[2]).strip()
        medio = row[3].capitalize()
        monto = float(row[4] or 0) + float(row[5] or 0)
        if nro not in boletas:
            boletas[nro] = {"total": 0, "detalle": []}
        boletas[nro]["total"] += monto
        boletas[nro]["detalle"].append(f"{medio} ${monto:,.0f}")

    # Insertar bloque resumen al final
    ws.append([])
    ws.append(["Resumen por Boleta"])
    titulo = ws[ws.max_row][0]
    titulo.font = Font(bold=True, size=14)
    titulo.fill = header_fill
    titulo.border = thin_border

    # Encabezados
    ws.append(["Nº Interno", "Total Boleta", "Detalle"])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.border = thin_border

    # Filas de boletas
    for nro, datos in boletas.items():
        ws.append([
            nro,
            datos["total"],
            " + ".join(datos["detalle"])
        ])
        for c in ws[ws.max_row]:
            c.border = thin_border
            if isinstance(c.value, (int, float)):
                c.number_format = '"$"#,##0'

    _autoajustar_columnas(ws)


# -------------- INICIALIZAR XLSX --------------
def inicializar_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        wb.remove(wb.active)

        ws = wb.create_sheet("planilla transacciones")
        ws.append(["Fecha", "Código Autorización Tarjetas", "Nº Interno Software",
                   "Medio de Pago", "Monto sin Propina", "Propina", "Total con Propina"])

        ws = wb.create_sheet("planilla repartos")
        ws.append(["Fecha", "Repartidor", "Dirección", "Monto", "Piso Empresa"])

        ws = wb.create_sheet("planilla egresos")
        ws.append(["Fecha", "Motivo", "Valor", "Nº Boleta/Factura"])

        ws = wb.create_sheet("planilla mermas")
        ws.append(["Fecha", "Motivo", "Valor"])

        ws = wb.create_sheet("planilla desgloses")
        ws.append(["Fecha", "Denominación", "Cantidad", "Total", "Tipo"])

        ws = wb.create_sheet("parametros")
        ws.append(["Parametro", "Valor"])

        wb.save(EXCEL_FILE)
    else:
        wb = load_workbook(EXCEL_FILE)
        for hoja in HOJAS_NECESARIAS:
            if hoja not in wb.sheetnames:
                ws = wb.create_sheet(hoja)
                if hoja == "planilla transacciones":
                    ws.append(["Fecha", "Código Autorización Tarjetas", "Nº Interno Software",
                               "Medio de Pago", "Monto sin Propina", "Propina", "Total con Propina"])
                elif hoja == "planilla repartos":
                    ws.append(["Fecha", "Repartidor", "Dirección", "Monto", "Piso Empresa"])
                elif hoja == "planilla egresos":
                    ws.append(["Fecha", "Motivo", "Valor", "Nº Boleta/Factura"])
                elif hoja == "planilla mermas":
                    ws.append(["Fecha", "Motivo", "Valor"])
                elif hoja == "planilla desgloses":
                    ws.append(["Fecha", "Denominación", "Cantidad", "Total", "Tipo"])
                elif hoja == "planilla cortesias":
                    ws.append(["Fecha", "Monto", "Motivo"])
                elif hoja == "parametros":
                    ws.append(["Parametro", "Valor"])
            
        wb.save(EXCEL_FILE)

# -------------- CAJA INICIAL --------------
def obtener_caja_inicial():
    if not os.path.exists(EXCEL_FILE):
        return None
    wb = load_workbook(EXCEL_FILE)
    if "parametros" not in wb.sheetnames:
        return None
    ws = wb["parametros"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == "caja_inicial":
            return row[1]
    return None

# --------- Iniciar Turno (Caja Inicial) ---------
@app.route("/iniciar_turno", methods=["POST"])
def iniciar_turno():
    inicializar_excel()
    cajero = request.form.get("cajero")
    turno = request.form.get("turno")
    valor = float(request.form.get("caja_inicial", 0))

    if not cajero or not turno or valor <= 0:
        flash("⚠️ Debes ingresar Cajero, Turno y Caja Inicial.")
        return redirect(url_for("index"))

    wb = load_workbook(EXCEL_FILE)
    ws = wb["parametros"]

    # Guardar parámetros
    def set_param(nombre, val):
        for row in ws.iter_rows(min_row=2):
            if row[0].value == nombre:
                row[1].value = val
                return
        ws.append([nombre, val])

    set_param("cajero", cajero)
    set_param("turno", turno)
    set_param("caja_inicial", valor)

    wb.save(EXCEL_FILE)

    # Guardar en sesión también
    session["cajero"] = cajero
    session["turno"] = turno
    session["caja_inicial"] = valor

    return redirect(url_for("index"))

@app.before_request
def exigir_turno_activo():
    rutas_protegidas = {
        "agregar_venta", "agregar_reparto", "agregar_egreso",
        "agregar_merma", "agregar_desglose", "agregar_cortesia",
        "cierre_caja"
    }
    if request.endpoint in rutas_protegidas:
        if not session.get("cajero") or not session.get("turno") or not session.get("caja_inicial"):
            flash("⚠️ Debes iniciar turno para realizar esta acción.", "danger")
            return redirect(url_for("index"))

# -------------- RESUMEN CAJA --------------
from openpyxl.styles import Alignment

def construir_resumen_caja(wb):
    thin_border, header_fill = _estilos_basicos()

    if "Resumen Caja" not in wb.sheetnames:
        ws_r = wb.create_sheet("Resumen Caja")
    else:
        ws_r = wb["Resumen Caja"]
        ws_r.delete_rows(1, ws_r.max_row)

    # Obtener cajero, turno y caja inicial
    cajero, turno, caja_inicial = "No registrado", "No registrado", 0
    if "parametros" in wb.sheetnames:
        ws_p = wb["parametros"]
        for row in ws_p.iter_rows(min_row=2, values_only=True):
            if row[0] == "cajero":
                cajero = row[1]
            elif row[0] == "turno":
                turno = row[1]
            elif row[0] == "caja_inicial":
                caja_inicial = row[1]

    # Estilo encabezado principal
    encabezados = [
        f"Cajero: {cajero}",
        f"Turno: {turno}",
        ws_r.append([f"Caja Inicial: ${caja_inicial if caja_inicial else 0:,.0f}"])
    ]

    fila = 1
    for texto in encabezados:
        celda = ws_r[f"A{fila}"]
        celda.value = texto
        celda.font = Font(bold=True, size=14)
        celda.fill = header_fill              # fondo gris
        celda.alignment = Alignment(horizontal="left", vertical="center")
        celda.border = thin_border
        fila += 1

    ws_r["A5"] = f"Resumen Caja - Exportado el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_r["A5"].font = Font(bold=True, size=12)



    # -------- Leer transacciones para totales --------
    pagos_total = {m: 0 for m in MEDIOS_VALIDOS}        # totales (con propina)
    tarjetas_sin_propina = 0                            # débito + crédito + prepago (sin propina)
    propinas_total = {m: 0 for m in MEDIOS_VALIDOS}  # nuevo diccionario


    if "planilla transacciones" in wb.sheetnames:
        ws_t = wb["planilla transacciones"]
        for row in ws_t.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7:
                continue
            medio = str(row[3]).lower().strip() if row[3] else ""
            monto_sin = float(row[4] or 0)
            propina = float(row[5] or 0)
            total = float(row[6] or 0)

            if medio in pagos_total:
                pagos_total[medio] += total
                propinas_total[medio] += propina

            if medio in ("debito", "credito", "prepago"):
                tarjetas_sin_propina += monto_sin

    # -------- DESGLOSE DE VENTAS --------
    ws_r.append([])
    ws_r.append(["Desglose de Ventas"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
    for medio in MEDIOS_VALIDOS:
        ws_r.append([medio.capitalize(), pagos_total.get(medio, 0)])
        ws_r[f"B{ws_r.max_row}"].number_format = '"$"#,##0'
        for c in ws_r[ws_r.max_row]:
            c.border = thin_border

    # -------- RESUMEN TARJETAS --------
    ws_r.append([])
    ws_r.append(["Resumen Tarjetas"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
    total_tarjetas_con_prop = pagos_total.get("debito",0) + pagos_total.get("credito",0) + pagos_total.get("prepago",0)
    ws_r.append(["Débito", pagos_total.get("debito",0)])
    ws_r.append(["Crédito", pagos_total.get("credito",0)])
    ws_r.append(["Prepago", pagos_total.get("prepago",0)])
    ws_r.append(["Total sin Propinas", tarjetas_sin_propina])
    ws_r.append(["Total con Propinas", total_tarjetas_con_prop])
    for row in ws_r.iter_rows(min_row=ws_r.max_row-4, max_row=ws_r.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '"$"#,##0'
            cell.border = thin_border

    # -------- RESUMEN EFECTIVO --------
    ws_r.append([])
    ws_r.append(["Resumen Efectivo"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
    caja_inicial = obtener_caja_inicial()
    if caja_inicial is None:
            caja_inicial = 0
    venta_efectivo = pagos_total.get("efectivo", 0)
    egresos_ef = 0
    if "planilla egresos" in wb.sheetnames:
        ws_e = wb["planilla egresos"]
        for r in ws_e.iter_rows(min_row=2, values_only=True):
            if r and len(r) >= 3:
                egresos_ef += float(r[2] or 0)
    total_efectivo = (caja_inicial + venta_efectivo) - egresos_ef
    ws_r.append(["Caja Inicial", caja_inicial])
    ws_r.append(["Venta Efectivo", venta_efectivo])
    ws_r.append(["Total Egresos Efectivo", egresos_ef])
    ws_r.append(["Total Resumen Efectivo", total_efectivo])
    for row in ws_r.iter_rows(min_row=ws_r.max_row-3, max_row=ws_r.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '"$"#,##0'
            cell.border = thin_border

    # -------- PROPINAS --------
# -------- PROPINAS DETALLADAS --------
    ws_r.append([])
    ws_r.append(["Propinas por Medio de Pago"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)

    total_propinas = 0
    for medio, valor in propinas_total.items():
     if valor > 0:
        ws_r.append([f"Propinas {medio.capitalize()}", valor])
        ws_r[f"B{ws_r.max_row}"].number_format = '"$"#,##0'
        for c in ws_r[ws_r.max_row]:
            c.border = thin_border
        total_propinas += valor

    ws_r.append(["Total Propinas", total_propinas])
    ws_r[f"B{ws_r.max_row}"].number_format = '"$"#,##0'
    for c in ws_r[ws_r.max_row]:
        c.font = Font(bold=True)
        c.border = thin_border


    # -------- DESGLOSES (Caja / Depositar) --------
    desg_caja = {d: 0 for d in DENOMINACIONES}
    desg_dep = {d: 0 for d in DENOMINACIONES}
    if "planilla desgloses" in wb.sheetnames:
        ws_d = wb["planilla desgloses"]
        for r in ws_d.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            denom = int(r[1] or 0)
            cant = int(r[2] or 0)
            tipo = r[4] or "Caja"
            if denom in DENOMINACIONES and cant > 0:
                if str(tipo).lower() == "caja":
                    desg_caja[denom] += cant
                else:
                    desg_dep[denom] += cant

    # - Caja
    ws_r.append([])
    ws_r.append(["Desglose Caja"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
    ws_r.append(["Denominación", "Cantidad", "Total"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
    total_caja = 0
    for d in DENOMINACIONES:
        if desg_caja[d] > 0:
            total = d * desg_caja[d]
            total_caja += total
            ws_r.append([f"${d:,}", desg_caja[d], total])
            ws_r[f"C{ws_r.max_row}"].number_format = '"$"#,##0'
            for c in ws_r[ws_r.max_row]:
                c.border = thin_border
    ws_r.append(["Total Caja", "", total_caja])
    ws_r[f"C{ws_r.max_row}"].number_format = '"$"#,##0'

    # - Depositar
    ws_r.append([])
    ws_r.append(["Desglose Efectivo a Depositar"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
    ws_r.append(["Denominación", "Cantidad", "Total"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
    total_dep = 0
    for d in DENOMINACIONES:
        if desg_dep[d] > 0:
            total = d * desg_dep[d]
            total_dep += total
            ws_r.append([f"${d:,}", desg_dep[d], total])
            ws_r[f"C{ws_r.max_row}"].number_format = '"$"#,##0'
            for c in ws_r[ws_r.max_row]:
                c.border = thin_border
    ws_r.append(["Total a Depositar", "", total_dep])
    ws_r[f"C{ws_r.max_row}"].number_format = '"$"#,##0'

 # -------- REPARTOS (por repartidor + piso) --------
    ws_r.append([])
    ws_r.append(["Repartos"])
    _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)

    repartidores = {}  # nombre -> {total_montos, piso}
    if "planilla repartos" in wb.sheetnames:
        ws_rep = wb["planilla repartos"]
        for r in ws_rep.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            nombre = (r[1] or "").strip()
            monto = float(r[3] or 0)
            piso = float(r[4] or 0)
            if not nombre:
                continue
            if nombre not in repartidores:
                repartidores[nombre] = {"total": 0.0, "piso": piso}
            repartidores[nombre]["total"] += monto
            # mantener el primer piso que se haya ingresado
            if repartidores[nombre]["piso"] == 0 and piso > 0:
                repartidores[nombre]["piso"] = piso

    if repartidores:
        ws_r.append(["Repartidor", "Total Repartos", "Piso Empresa", "Total Final"])
        _estilizar_encabezado(ws_r[ws_r.max_row], header_fill, thin_border)
        for nombre, datos in repartidores.items():
            total_final = datos["total"] + datos["piso"]
            ws_r.append([nombre, datos["total"], datos["piso"], total_final])
            for c in ws_r[ws_r.max_row]:
                if isinstance(c.value, (int, float)):
                    c.number_format = '"$"#,##0'
                c.border = thin_border

    # -------- TOTAL CAJA (ORDENADO Y DETALLADO) --------
    ws_r.append([])
    ws_r.append(["RESUMEN DE CAJA"])
    ws_r["A" + str(ws_r.max_row)].font = Font(bold=True, size=16)
    ws_r["A" + str(ws_r.max_row)].fill = header_fill
    ws_r["A" + str(ws_r.max_row)].alignment = Alignment(horizontal="center")

    # Calcular totales
    caja_inicial = obtener_caja_inicial() or 0
    total_ventas = sum(pagos_total.values())

    total_egresos = 0
    total_cortesias = 0
    total_mermas = 0

    if "planilla egresos" in wb.sheetnames:
        ws_e = wb["planilla egresos"]
        total_egresos = sum(float(r[2] or 0) for r in ws_e.iter_rows(min_row=2, values_only=True))

    if "planilla cortesias" in wb.sheetnames:
        ws_c = wb["planilla cortesias"]
        total_cortesias = sum(float(r[1] or 0) for r in ws_c.iter_rows(min_row=2, values_only=True))

    if "planilla mermas" in wb.sheetnames:
        ws_m = wb["planilla mermas"]
        total_mermas = sum(float(r[2] or 0) for r in ws_m.iter_rows(min_row=2, values_only=True))

    total_caja = caja_inicial + total_ventas - total_egresos - total_cortesias - total_mermas

    # Calcular porcentaje de pérdidas (cortesías + mermas)
    porcentaje_perdidas = 0
    if total_ventas > 0:
        porcentaje_perdidas = ((total_cortesias + total_mermas) / total_ventas) * 100

    # Mostrar totales organizados en dos columnas
    resumen_datos = [
        ("Caja Inicial", caja_inicial),
        ("Ventas Totales", total_ventas),
        ("Egresos", -total_egresos),
        ("Cortesías", -total_cortesias),
        ("Mermas", -total_mermas),
        ("TOTAL CAJA FINAL", total_caja),
        ("", ""),  # Espacio visual
        ("% Pérdidas sobre Ventas", f"{porcentaje_perdidas:.2f}%"),
    ]

    ws_r.append([])
    for label, valor in resumen_datos:
        ws_r.append([label, valor])

    # Aplicar formato visual
    for fila in range(ws_r.max_row - len(resumen_datos) + 1, ws_r.max_row + 1):
        c1 = ws_r[f"A{fila}"]
        c2 = ws_r[f"B{fila}"]

        c1.font = Font(bold=True if "TOTAL" in str(c1.value) else False, size=13)
        c2.font = Font(bold=True if "TOTAL" in str(c1.value) else False, size=13)

        if isinstance(c2.value, (int, float)):
            c2.number_format = '"$"#,##0'
        c1.alignment = Alignment(horizontal="left")
        c2.alignment = Alignment(horizontal="right")
        c1.border = Border(bottom=Side(style="thin"))
        c2.border = Border(bottom=Side(style="thin"))

    _autoajustar_columnas(ws_r)


def estilizar_hojas_detalle(wb):
    thin_border, header_fill = _estilos_basicos()
    for nombre in wb.sheetnames:
        if nombre == "Resumen Caja":
            continue
        _estilizar_hoja_detalle(wb[nombre], header_fill, thin_border)

# ---------------- RUTAS UI ----------------
@app.route("/")
def index():
    return render_template("index.html")

# Registrar venta (total con propina se calcula solo)
@app.route("/agregar_venta", methods=["GET","POST"])
def agregar_venta():
    if request.method == "POST":
        inicializar_excel()
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        numero_interno = request.form.get("numero_interno", "")
        codigo_autorizacion = request.form.get("codigo_autorizacion", "")

        # Recibir múltiples pagos y propinas
        medios = request.form.getlist("medio_pago[]")
        montos = [float(m or 0) for m in request.form.getlist("monto_pago[]")]
        propinas = [float(p or 0) for p in request.form.getlist("propina_pago[]")]

        wb = load_workbook(EXCEL_FILE)
        ws = wb["planilla transacciones"]

        total_boleta = 0
        for medio, monto, propina in zip(medios, montos, propinas):
            total = monto + propina
            total_boleta += total
            ws.append([
                fecha,
                codigo_autorizacion if medio.lower() in ("debito", "credito") else "",
                numero_interno,
                medio.lower(),
                monto,
                propina,
                total
            ])

        wb.save(EXCEL_FILE)

        return render_template(
            "result.html",
            mensaje=f"✅ Venta registrada Nº {numero_interno}: ${total_boleta:,.0f}",
            volver="agregar_venta"
        )

    return render_template("agregar_venta.html", medios=MEDIOS_VALIDOS)



# Registrar reparto (con Piso Empresa, admite 0/5000/10000)
@app.route("/agregar_reparto", methods=["GET","POST"])
def agregar_reparto():
    if request.method == "POST":
        inicializar_excel()
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        repartidor = request.form.get("repartidor", "").strip()
        direccion = request.form.get("direccion", "")
        monto = float(request.form.get("monto", 0))
        piso = float(request.form.get("piso") or 0)


        wb = load_workbook(EXCEL_FILE)
        ws = wb["planilla repartos"]

        # --- validar piso existente ---
        piso_existente = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and r[1] and r[1].strip().lower() == repartidor.lower():
                piso_existente = piso_existente or float(r[4] or 0)

        if piso_existente > 0:
            piso = 0   # Si ya tenía piso, este se ignora

        ws.append([fecha, repartidor, direccion, monto, piso])
        wb.save(EXCEL_FILE)
        return render_template("result.html", mensaje="🚚 Reparto registrado con éxito", volver="agregar_reparto")
    return render_template("agregar_reparto.html")


# Registrar egreso
@app.route("/agregar_egreso", methods=["GET","POST"])
def agregar_egreso():
    if request.method == "POST":
        inicializar_excel()
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb = load_workbook(EXCEL_FILE)
        ws = wb["planilla egresos"]
        ws.append([
            fecha,
            request.form.get("motivo", ""),
            float(request.form.get("valor", 0)),
            request.form.get("boleta", "")
        ])
        wb.save(EXCEL_FILE)
        return render_template("result.html", mensaje="💸 Egreso registrado con éxito", volver="agregar_egreso")
    return render_template("agregar_egreso.html")

# Registrar merma
@app.route("/agregar_merma", methods=["GET","POST"])
def agregar_merma():
    if request.method == "POST":
        inicializar_excel()
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb = load_workbook(EXCEL_FILE)
        ws = wb["planilla mermas"]
        ws.append([
            fecha,
            request.form.get("motivo", ""),
            float(request.form.get("valor", 0))
        ])
        wb.save(EXCEL_FILE)
        return render_template("result.html", mensaje="⚠️ Merma registrada con éxito", volver="agregar_merma")
    return render_template("agregar_merma.html")

# Registrar desglose (total calculado automáticamente)
@app.route("/agregar_desglose", methods=["GET", "POST"])
def agregar_desglose():
    if request.method == "POST":
        try:
            inicializar_excel()
            fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # --- Limpieza segura de datos del formulario ---
            den_str = request.form.get("denominacion", "0").replace("$", "").replace(".", "").replace(",", "").strip()
            cant_str = request.form.get("cantidad", "0").replace(",", "").strip()

            # Validar que sean numéricos
            try:
                den = int(den_str)
                cant = int(cant_str)
            except ValueError:
                flash("⚠️ La denominación y cantidad deben ser números válidos.", "warning")
                return render_template("agregar_desglose.html", denominaciones=DENOMINACIONES)

            total = den * cant
            tipo = request.form.get("tipo", "Caja")

            wb = load_workbook(EXCEL_FILE)

            # Crear hoja si no existe
            if "planilla desgloses" not in wb.sheetnames:
                ws = wb.create_sheet("planilla desgloses")
                ws.append(["Fecha", "Denominación", "Cantidad", "Total", "Tipo"])
            else:
                ws = wb["planilla desgloses"]

            ws.append([fecha, den, cant, total, tipo])
            wb.save(EXCEL_FILE)

            return render_template(
                "result.html",
                mensaje="💵 Desglose registrado con éxito",
                volver="agregar_desglose"
            )

        except Exception as e:
            print(f"❌ Error en agregar_desglose: {e}")
            flash("Ocurrió un error inesperado al agregar el desglose.", "danger")
            return render_template("agregar_desglose.html", denominaciones=DENOMINACIONES)

    return render_template("agregar_desglose.html", denominaciones=DENOMINACIONES)

@app.route("/agregar_cortesia", methods=["GET", "POST"])
def agregar_cortesia():
    # Bloquear si no hay cajero logueado
    if not session.get("cajero"):
        flash("⚠️ Debes iniciar un turno para registrar cortesías.", "warning")
        return redirect(url_for("index"))

    if request.method == "POST":
        monto = request.form.get("monto")
        motivo = request.form.get("motivo")

        if not monto or not motivo:
            flash("⚠️ Debes ingresar monto y motivo.", "danger")
            return redirect(url_for("agregar_cortesia"))

        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb["planilla cortesias"]

            # Asegurar encabezado correcto
            if ws.max_row == 0:
                ws.append(["Fecha", "Monto", "Motivo"])
            else:
                encabezado = [cell.value for cell in ws[1]]
                if encabezado != ["Fecha", "Monto", "Motivo"]:
                    ws.delete_rows(1, ws.max_row)  # borrar todo
                    ws.append(["Fecha", "Monto", "Motivo"])

            # Agregar la nueva fila
            fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([fecha, int(monto), motivo])

            wb.save(EXCEL_FILE)

            # Mostrar pantalla de éxito (similar a desglose)
            return render_template(
                "result.html",
                mensaje=f"🎁 Cortesía registrada con éxito: ${monto} - {motivo}",
                volver="agregar_cortesia"
            )

        except Exception as e:
            flash(f"❌ Error al guardar cortesía: {e}", "danger")
            return redirect(url_for("agregar_cortesia"))

    return render_template("agregar_cortesia.html")

# Vistas simples de planillas
@app.route("/planilla_caja")
def planilla_caja():
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["planilla transacciones"]
    ventas = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    return render_template("planilla_caja.html", ventas=ventas)
# ------------------- ELIMINAR VENTA CON MOTIVO -------------------
@app.route("/eliminar_venta/<int:indice>", methods=["POST"])
def eliminar_venta(indice):
    """Elimina una venta solo si la clave es correcta, guarda el motivo y registra la venta borrada."""
    clave = request.form.get("clave_eliminar", "").strip()
    motivo = request.form.get("motivo_eliminar", "").strip()
    CLAVE_PERMITIDA = "frayesgustitos2025"

    if clave != CLAVE_PERMITIDA:
        flash("❌ Clave incorrecta. No se eliminó la venta.", "error")
        return redirect(url_for("planilla_caja"))

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["planilla transacciones"]

        # Extraer los valores de la fila antes de borrarla
        if indice >= 2 and indice <= ws.max_row:
            fila = [cell.value for cell in ws[indice]]
            ws.delete_rows(indice)
            wb.save(EXCEL_FILE)

            # Registrar venta borrada
            nombre_tabla_borradas = "Ventas Borradas"
            if nombre_tabla_borradas not in wb.sheetnames:
                ws_borradas = wb.create_sheet(nombre_tabla_borradas)
                ws_borradas.append(["Fecha Eliminación", "Código Autorización", "N° Interno", "Medio Pago", "Monto", "Propina", "Total", "Motivo"])
            else:
                ws_borradas = wb[nombre_tabla_borradas]

            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws_borradas.append([
                fecha_actual,
                fila[1] if len(fila) > 1 else "-",
                fila[2] if len(fila) > 2 else "-",
                fila[3] if len(fila) > 3 else "-",
                fila[4] if len(fila) > 4 else 0,
                fila[5] if len(fila) > 5 else 0,
                fila[6] if len(fila) > 6 else 0,
                motivo or "(sin motivo)"
            ])

            wb.save(EXCEL_FILE)
            flash("🗑️ Venta eliminada y registrada en 'Ventas Borradas'.", "success")

        else:
            flash("⚠️ No se pudo eliminar la venta (índice fuera de rango).", "error")

    except Exception as e:
        flash(f"Error al eliminar venta: {str(e)}", "error")

    return redirect(url_for("planilla_caja"))



@app.route("/planilla_repartos")
def planilla_repartos():
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["planilla repartos"]
    repartos = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    return render_template("planilla_repartos.html", repartos=repartos)

# ------------------- PLANILLA EGRESOS -------------------
@app.route("/planilla_egresos")
def planilla_egresos():
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["planilla egresos"]
    egresos = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    return render_template("planilla_egresos.html", egresos=egresos)

# ------------------- EDITAR EGRESO -------------------
@app.route("/editar_egreso/<int:indice>", methods=["GET", "POST"])
def editar_egreso(indice):
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb["planilla egresos"]

    # Validar índice (fila en Excel, incluye encabezado en la fila 1)
    if indice < 2 or indice > ws.max_row:
        flash("⚠️ No se encontró el egreso a editar (índice fuera de rango).", "error")
        return redirect(url_for("planilla_egresos"))

    if request.method == "POST":
        try:
            motivo = request.form.get("motivo", "").strip()
            valor_raw = request.form.get("valor", "0").replace(".", "").replace(",", "").strip()
            boleta = request.form.get("boleta", "").strip()

            try:
                valor = float(valor_raw or 0)
            except ValueError:
                flash("⚠️ El valor del egreso debe ser numérico.", "error")
                return redirect(url_for("editar_egreso", indice=indice))

            # Actualizar celdas (columna 1 = Fecha, 2 = Motivo, 3 = Valor, 4 = Nº Boleta/Factura)
            ws.cell(row=indice, column=2).value = motivo
            ws.cell(row=indice, column=3).value = valor
            ws.cell(row=indice, column=4).value = boleta

            wb.save(EXCEL_FILE)
            flash("✅ Egreso actualizado correctamente.", "success")
            return redirect(url_for("planilla_egresos"))

        except Exception as e:
            flash(f"❌ Error al actualizar el egreso: {e}", "error")
            return redirect(url_for("planilla_egresos"))

    # GET: cargar datos actuales
    fila = [c.value for c in ws[indice]]
    egreso = {
        "fecha": fila[0],
        "motivo": fila[1],
        "valor": fila[2],
        "boleta": fila[3],
    }

    return render_template("editar_egreso.html", indice=indice, egreso=egreso)


# ------------------- ELIMINAR EGRESO -------------------
@app.route("/eliminar_egreso/<int:indice>", methods=["POST"])
def eliminar_egreso(indice):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["planilla egresos"]

        if indice >= 2 and indice <= ws.max_row:
            ws.delete_rows(indice)
            wb.save(EXCEL_FILE)
            flash("🗑️ Egreso eliminado correctamente.", "success")
        else:
            flash("⚠️ No se pudo eliminar el egreso (índice fuera de rango).", "error")

    except Exception as e:
        flash(f"❌ Error al eliminar egreso: {e}", "error")

    return redirect(url_for("planilla_egresos"))


# ------------------- ELIMINAR REPARTO -------------------
@app.route("/eliminar_reparto/<int:indice>", methods=["POST"])
def eliminar_reparto(indice):
    """Elimina un reparto de la planilla repartos según su índice."""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["planilla repartos"]

        if indice >= 2 and indice <= ws.max_row:
            ws.delete_rows(indice)
            wb.save(EXCEL_FILE)
            flash("🗑️ Reparto eliminado correctamente.", "success")
        else:
            flash("⚠️ No se pudo eliminar el reparto (índice fuera de rango).", "error")

    except Exception as e:
        flash(f"Error al eliminar reparto: {str(e)}", "error")

    return redirect(url_for("planilla_repartos"))


# --------- Descargar Excel actual (con Resumen Caja y estilos) ---------
@app.route("/descargar_actual")
def descargar_actual():
    inicializar_excel()
    wb = load_workbook(EXCEL_FILE)
    construir_resumen_caja(wb)
    estilizar_hojas_detalle(wb)
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio, as_attachment=True,
        download_name=f"planilla_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# --------- Cierre de Caja (guarda, limpia y resetea Caja Inicial) ---------
# --------- Cierre de Caja ---------
# --------- Cierre de Caja (guarda, limpia y resetea Caja Inicial) ---------
@app.route("/cierre_caja")
def cierre_caja():
    # 🔒 Verificar que haya turno activo
    if not session.get("cajero") or not session.get("turno") or not session.get("caja_inicial"):
        flash("⚠️ No puedes cerrar caja sin haber iniciado un turno.", "danger")
        return redirect(url_for("index"))

    inicializar_excel()
    wb = load_workbook(EXCEL_FILE)

    # Construir resumen y aplicar estilos
    construir_resumen_caja(wb)
    estilizar_hojas_detalle(wb)

    # Generar bloque resumen por boleta
    resumen_boletas_en_transacciones(wb)

    # Guardar archivo de cierre
    nombre = f"Cierre caja {datetime.now().strftime('%d-%m-%Y_%H-%M-%S')} Camilo Henriquez.xlsx"

    ruta = os.path.join(CIERRES_DIR, nombre)
    wb.save(ruta)

    # Limpiar planillas para el nuevo turno
    for hoja in [
        "planilla transacciones", "planilla repartos", "planilla egresos",
        "planilla mermas", "planilla desgloses", "planilla cortesias"
    ]:
        if hoja in wb.sheetnames:
            ws = wb[hoja]
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row)

            # Asegurar encabezado correcto en cortesías
            if hoja == "planilla cortesias":
                if ws.max_row == 0:
                    ws.append(["Fecha", "Monto", "Motivo"])
                else:
                    encabezado = [cell.value for cell in ws[1]]
                    if encabezado != ["Fecha", "Monto", "Motivo"]:
                        ws.delete_rows(1, ws.max_row)
                        ws.append(["Fecha", "Monto", "Motivo"])

    # Limpiar hoja Resumen Caja
    if "Resumen Caja" in wb.sheetnames:
        ws = wb["Resumen Caja"]
        ws.delete_rows(1, ws.max_row)
    
     # Limpiar hoja Ventas Borradas (pero manteniendo el encabezado)
    if "Ventas Borradas" in wb.sheetnames:
        ws_vb = wb["Ventas Borradas"]
        if ws_vb.max_row > 1:
            ws_vb.delete_rows(2, ws_vb.max_row)

    # Resetear parámetros
    if "parametros" in wb.sheetnames:
        ws = wb["parametros"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "caja_inicial":
                row[1].value = 0
            elif row[0].value in ("cajero", "turno"):
                row[1].value = None

    wb.save(EXCEL_FILE)

    # Guardar archivo de cierre en sesión
    session["archivo_cierre"] = ruta

    # 🔹 Limpiar cajero/turno de la sesión (pero no borrar archivo_cierre)
    session.pop("cajero", None)
    session.pop("turno", None)
    session.pop("caja_inicial", None)

    # Redirigir a pantalla de Caja Cerrada
    return redirect(url_for("caja_cerrada"))


# --------- Pantalla de Caja Cerrada ---------
@app.route("/caja_cerrada")
def caja_cerrada():
    return render_template("caja_cerrada.html")




# --------- Historial de cierres ---------
@app.route("/historial_cierres")
def historial_cierres():
    archivos = [f for f in os.listdir(CIERRES_DIR) if f.lower().endswith(".xlsx")]
    archivos.sort(reverse=True)
    return render_template("historial_cierres.html", archivos=archivos)

@app.route("/descargar_cierre/<nombre>")
def descargar_cierre(nombre):
    ruta = os.path.join(CIERRES_DIR, nombre)
    if os.path.exists(ruta):
        return send_file(
            ruta, as_attachment=True, download_name=nombre,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    return "Archivo no encontrado", 404


# --------- Descargar archivo de cierre ---------
@app.route("/descargar_cierre_final")
def descargar_cierre_final():
    try:
        archivo = session.get("archivo_cierre")
        print("📦 Archivo en sesión:", archivo)

        # Si no existe en sesión, buscar el más reciente en la carpeta
        if not archivo or not os.path.exists(archivo):
            print("⚠️ Buscando el cierre más reciente en carpeta...")
            archivos = sorted(
                [os.path.join(CIERRES_DIR, f) for f in os.listdir(CIERRES_DIR) if f.endswith(".xlsx")],
                key=os.path.getmtime,
                reverse=True
            )
            if archivos:
                archivo = archivos[0]
                print("✅ Usando más reciente:", archivo)
            else:
                flash("⚠️ No se encontró ningún archivo de cierre para descargar.", "danger")
                return redirect(url_for("index"))

        if not os.path.exists(archivo):
            flash("❌ El archivo de cierre no existe o fue movido.", "danger")
            return redirect(url_for("index"))

        nombre = os.path.basename(archivo)
        print("🔽 Descargando:", nombre)

        return send_file(
            archivo,
            as_attachment=True,
            download_name=nombre,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"❌ Error inesperado en descargar_cierre_final: {e}")
        flash("Ocurrió un error al descargar el archivo de cierre.", "danger")
        return redirect(url_for("index"))


# ---------------- MAIN ----------------
if __name__ == "__main__":
    import threading, webbrowser
    from flask import request
    import atexit

    def open_browser():
        webbrowser.open("http://127.0.0.1:5000")

    # Función para cerrar Flask cuando se cierre la app
    def shutdown_server():
        os.kill(os.getpid(), signal.SIGTERM)

    @app.route('/shutdown', methods=['POST'])
    def shutdown():
        shutdown_server()
        return "Servidor detenido."

    # Registrar cierre al salir
    atexit.register(shutdown_server)

    inicializar_excel()
    threading.Timer(1, open_browser).start()
    app.run(host="127.0.0.1", port=5000, debug=False)

